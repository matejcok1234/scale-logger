import tkinter as tk
from tkinter import ttk, messagebox
import threading
import serial
import serial.tools.list_ports
import datetime
import csv
import os
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

SAMPLE_THRESHOLD = 70.0  # grams — above this = sample on scale (record it)
EMPTY_THRESHOLD  = 10.0  # grams — below this = scale empty (ready for next reading)
POLL_INTERVAL    = 0.5   # seconds between polls

# Carousel layout — must match the HT03RA100 firmware: 8 samples, each measured
# 5 times (the probe dips onto each sample 5×). Every 5 detected weigh-events
# therefore belong to ONE physical sample; samples are numbered 1..8. The run
# repeats the whole carousel indefinitely (until the user presses Stop), so a
# "cycle" counter increments each time all 8 samples have been measured 5× again.
N_SAMPLES        = 8
MEAS_PER_SAMPLE  = 5
SET_SIZE         = N_SAMPLES * MEAS_PER_SAMPLE   # 40 readings = one full cycle

BG = '#f4f4f4'
BLUE  = '#1565C0'
GREEN = '#2e7d32'
RED   = '#c62828'


class ScaleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Scale Logger")
        self.root.configure(bg=BG)
        self.root.resizable(False, False)

        self.serial_conn  = None
        self.serial_lock  = threading.Lock()
        self.running      = False
        self.state        = 'WAITING'   # WAITING | SETTLING | RECORDED
        self.meas_count   = 0           # total measurements recorded this run (grows until Stop)
        self.results      = []

        self._build_ui()
        self._refresh_ports()

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        p = dict(padx=12, pady=6)

        # Port row
        pf = tk.Frame(self.root, bg=BG)
        pf.grid(row=0, column=0, sticky='ew', **p)

        tk.Label(pf, text="Port:", bg=BG).pack(side='left')
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(pf, textvariable=self.port_var, width=20, state='readonly')
        self.port_combo.pack(side='left', padx=(4, 2))
        tk.Button(pf, text="↻", command=self._refresh_ports, width=2).pack(side='left', padx=2)
        self.btn_connect = tk.Button(pf, text="Connect", command=self._toggle_connect, width=10)
        self.btn_connect.pack(side='left', padx=6)
        self.btn_tare = tk.Button(pf, text="Tare", command=self._tare, width=8, state='disabled')
        self.btn_tare.pack(side='left')

        # Live weight display
        wf = tk.LabelFrame(self.root, text="Live Weight", bg=BG, font=('Arial', 9))
        wf.grid(row=1, column=0, sticky='ew', **p)

        self.weight_var = tk.StringVar(value="---")
        tk.Label(wf, textvariable=self.weight_var, font=('Courier', 48, 'bold'),
                 bg=BG, fg='#111', width=10, anchor='e').pack(side='left', padx=(10, 0))
        tk.Label(wf, text="g", font=('Arial', 20), bg=BG, fg='#555').pack(side='left', anchor='s', pady=14)

        self.stability_var = tk.StringVar(value="")
        self.stability_lbl = tk.Label(wf, textvariable=self.stability_var,
                                      font=('Arial', 12, 'bold'), bg=BG, width=14, anchor='w')
        self.stability_lbl.pack(side='left', padx=16)

        # Sample names
        nf = tk.LabelFrame(self.root, text="Sample Names  (8 samples, each measured 5× — fill in before starting)", bg=BG, font=('Arial', 9))
        nf.grid(row=2, column=0, sticky='ew', **p)

        self.name_vars = []
        for i in range(N_SAMPLES):
            r, c = divmod(i, 2)
            var = tk.StringVar(value=f"Sample {i+1}")
            self.name_vars.append(var)
            tk.Label(nf, text=f"{i+1}:", bg=BG, width=2, anchor='e').grid(
                row=r, column=c*2, sticky='e', padx=(10, 2), pady=2)
            tk.Entry(nf, textvariable=var, width=20).grid(
                row=r, column=c*2+1, sticky='w', padx=(0, 14), pady=2)

        # Control buttons
        cf = tk.Frame(self.root, bg=BG)
        cf.grid(row=3, column=0, sticky='ew', **p)

        self.btn_start = tk.Button(
            cf, text="▶  Start Run", command=self._start_run,
            width=16, state='disabled',
            bg=GREEN, fg='white', font=('Arial', 12, 'bold'),
            relief='flat', cursor='hand2', pady=6)
        self.btn_start.pack(side='left', padx=(0, 10))

        self.btn_stop = tk.Button(
            cf, text="■  Stop & Export", command=self._stop_run,
            width=16, state='disabled',
            bg=RED, fg='white', font=('Arial', 12, 'bold'),
            relief='flat', cursor='hand2', pady=6)
        self.btn_stop.pack(side='left')

        # Status
        self.status_var = tk.StringVar(value="Not connected")
        tk.Label(self.root, textvariable=self.status_var, font=('Arial', 10),
                 bg=BG, fg='#444', anchor='w').grid(row=4, column=0, sticky='ew', padx=12)

        # Results table
        tf = tk.LabelFrame(self.root, text="Recorded Samples", bg=BG, font=('Arial', 9))
        tf.grid(row=5, column=0, sticky='nsew', **p)

        cols = ('cycle', 'sample', 'meas', 'name', 'weight', 'time')
        self.tree = ttk.Treeview(tf, columns=cols, show='headings', height=10)
        self.tree.heading('cycle',  text='Cycle')
        self.tree.heading('sample', text='Sample')
        self.tree.heading('meas',   text='Meas')
        self.tree.heading('name',   text='Name')
        self.tree.heading('weight', text='Weight')
        self.tree.heading('time',   text='Time')
        self.tree.column('cycle',  width=50,  anchor='center')
        self.tree.column('sample', width=55,  anchor='center')
        self.tree.column('meas',   width=50,  anchor='center')
        self.tree.column('name',   width=160)
        self.tree.column('weight', width=115, anchor='center')
        self.tree.column('time',   width=85,  anchor='center')

        sb = ttk.Scrollbar(tf, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        self.root.columnconfigure(0, weight=1)

    # ── Port management ────────────────────────────────────────────────────────

    def _refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.port_combo['values'] = ports
        if ports and not self.port_var.get():
            self.port_var.set(ports[0])

    def _toggle_connect(self):
        if self.serial_conn and self.serial_conn.is_open:
            self._disconnect()
        else:
            self._connect()

    def _connect(self):
        port = self.port_var.get()
        if not port:
            messagebox.showerror("Error", "Select a port first.")
            return
        try:
            self.serial_conn = serial.Serial(port, 9600, timeout=2)
            self.btn_connect.config(text="Disconnect")
            self.btn_tare.config(state='normal')
            self.btn_start.config(state='normal')
            self.status_var.set(f"Connected to {port}")
            threading.Thread(target=self._read_loop, daemon=True).start()
        except serial.SerialException as e:
            messagebox.showerror("Connection Error", str(e))

    def _disconnect(self):
        self.running = False
        conn, self.serial_conn = self.serial_conn, None
        if conn:
            try:
                conn.close()
            except Exception:
                pass
        self.btn_connect.config(text="Connect")
        self.btn_tare.config(state='disabled')
        self.btn_start.config(state='disabled')
        self.btn_stop.config(state='disabled')
        self.weight_var.set("---")
        self.stability_var.set("")
        self.status_var.set("Disconnected")

    # ── Scale commands ─────────────────────────────────────────────────────────

    def _tare(self):
        self._send(b'T\r\n')

    def _send(self, cmd):
        if self.serial_conn and self.serial_conn.is_open:
            with self.serial_lock:
                try:
                    self.serial_conn.write(cmd)
                except Exception:
                    pass

    # ── Run control ────────────────────────────────────────────────────────────

    def _next_cycle_num(self):
        # 1-based cycle (full pass through all 8 samples) for the next reading
        return self.meas_count // SET_SIZE + 1

    def _next_sample_num(self):
        # 1-based sample number (1..8) for the NEXT measurement, wrapping each cycle
        return self.meas_count % SET_SIZE // MEAS_PER_SAMPLE + 1

    def _next_meas_num(self):
        # 1-based measurement index (1..MEAS_PER_SAMPLE) within the current sample
        return self.meas_count % MEAS_PER_SAMPLE + 1

    def _start_run(self):
        self.results    = []
        self.meas_count = 0
        self.state      = 'WAITING'
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._send(b'T\r\n')        # auto-tare the scale at the start of the run
        self.running = True
        self.btn_start.config(state='disabled')
        self.btn_stop.config(state='normal')
        self.status_var.set(
            f"Tared — waiting for sample 1 · meas 1/{MEAS_PER_SAMPLE}  (runs until you press Stop)")

    def _stop_run(self):
        self.running = False
        self.btn_start.config(state='normal')
        self.btn_stop.config(state='disabled')
        n = len(self.results)
        self.status_var.set(f"Stopped — {n} measurement{'s' if n != 1 else ''} recorded")
        if self.results:
            self._export()

    # ── Serial read loop ───────────────────────────────────────────────────────

    def _read_loop(self):
        while self.serial_conn and self.serial_conn.is_open:
            try:
                with self.serial_lock:
                    self.serial_conn.write(b'SI\r\n')  # immediate reading (stable or dynamic)
                    raw = self.serial_conn.readline()
                line = raw.decode(errors='ignore').strip()
                if line:
                    self._process(line)
            except Exception:
                self.root.after(0, self._on_disconnect_error)
                break
            time.sleep(POLL_INTERVAL)

    def _on_disconnect_error(self):
        self.weight_var.set("---")
        self.stability_var.set("")
        self.status_var.set("Connection lost — reconnect the scale")
        self.btn_connect.config(text="Connect")
        self.btn_tare.config(state='disabled')
        self.btn_start.config(state='disabled')
        self.btn_stop.config(state='disabled')
        self.running = False

    def _process(self, line):
        # MT-SICS SI response: "S S  value unit" (stable) or "S D  value unit" (dynamic)
        parts = line.split()
        if len(parts) < 4 or parts[0] != 'S':
            return
        stable = (parts[1] == 'S')
        try:
            weight = float(parts[2])
            unit   = parts[3]
        except ValueError:
            return

        self.root.after(0, self._update_display, weight, unit, stable)

        if not self.running:
            return

        if self.state == 'WAITING':
            if weight > SAMPLE_THRESHOLD:
                if stable:
                    self._record(weight, unit)
                else:
                    self.state = 'SETTLING'
                    self.root.after(0, self.status_var.set,
                                    f"Sample {self._next_sample_num()}/{N_SAMPLES} · "
                                    f"meas {self._next_meas_num()}/{MEAS_PER_SAMPLE} — settling...")

        elif self.state == 'SETTLING':
            if weight < EMPTY_THRESHOLD:
                # weight dropped before stable — false alarm, reset
                self.state = 'WAITING'
                self.root.after(0, self.status_var.set,
                                f"Waiting for sample {self._next_sample_num()} · "
                                f"meas {self._next_meas_num()}/{MEAS_PER_SAMPLE}...")
            elif stable:
                self._record(weight, unit)

        elif self.state == 'RECORDED':
            if weight < EMPTY_THRESHOLD:
                self.state = 'WAITING'
                self.root.after(0, self.status_var.set,
                                f"Waiting for sample {self._next_sample_num()} · "
                                f"meas {self._next_meas_num()}/{MEAS_PER_SAMPLE}...")

    def _record(self, weight, unit):
        idx        = self.meas_count                      # 0-based measurement index
        cycle_num  = idx // SET_SIZE + 1                  # 1, 2, 3, ... (full carousel pass)
        within_set = idx % SET_SIZE                        # 0..39 within this cycle
        sample_idx = within_set // MEAS_PER_SAMPLE        # 0..N_SAMPLES-1
        sample_num = sample_idx + 1                       # 1..8  (never 0-based)
        meas_num   = within_set % MEAS_PER_SAMPLE + 1     # 1..MEAS_PER_SAMPLE
        name = self.name_vars[sample_idx].get()           # sample_idx is always 0..7
        now  = datetime.datetime.now()

        entry = {
            'cycle':  cycle_num,
            'sample': sample_num,
            'meas':   meas_num,
            'name':   name,
            'weight': weight,
            'unit':   unit,
            'date':   now.strftime('%Y-%m-%d'),
            'time':   now.strftime('%H:%M:%S'),
        }
        self.results.append(entry)        # every individual reading is kept (no averaging)
        self.meas_count += 1
        self.state = 'RECORDED'

        self.root.after(0, self._add_row, entry)
        self.root.after(0, self.status_var.set,
                        f"✓  cycle {cycle_num} · {name} · meas {meas_num}/{MEAS_PER_SAMPLE}: "
                        f"{weight:.3f} {unit}   —   lift / remove")

    # ── Display ────────────────────────────────────────────────────────────────

    def _update_display(self, weight, unit, stable):
        self.weight_var.set(f"{weight:>9.3f}")
        if stable:
            self.stability_lbl.config(text="● STABLE",    fg=GREEN)
        else:
            self.stability_lbl.config(text="◌ SETTLING…", fg='#e65100')

    def _add_row(self, entry):
        self.tree.insert('', 'end', values=(
            entry['cycle'],
            entry['sample'],
            f"{entry['meas']}/{MEAS_PER_SAMPLE}",
            entry['name'],
            f"{entry['weight']:.3f} {entry['unit']}",
            entry['time'],
        ))
        self.tree.yview_moveto(1)

    # ── Export ─────────────────────────────────────────────────────────────────

    def _export(self):
        ts      = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        base    = os.path.join(desktop, f"scale_run_{ts}")
        saved   = []

        # CSV
        csv_path = base + '.csv'
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['Cycle', 'Sample #', 'Measurement', 'Sample Name', 'Weight', 'Unit', 'Date', 'Time'])
            for r in self.results:
                w.writerow([r['cycle'], r['sample'], r['meas'], r['name'],
                            r['weight'], r['unit'], r['date'], r['time']])
        saved.append(os.path.basename(csv_path))

        # Excel
        if HAS_EXCEL:
            xlsx_path = base + '.xlsx'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            hf = Font(bold=True, color='FFFFFF')
            hb = PatternFill('solid', fgColor='1565C0')
            ac = Alignment(horizontal='center')

            headers = ['Cycle', 'Sample #', 'Measurement', 'Sample Name', 'Weight', 'Unit', 'Date', 'Time']
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font, cell.fill, cell.alignment = hf, hb, ac

            for ri, r in enumerate(self.results, 2):
                ws.cell(ri, 1, r['cycle'])
                ws.cell(ri, 2, r['sample'])
                ws.cell(ri, 3, r['meas'])
                ws.cell(ri, 4, r['name'])
                ws.cell(ri, 5, r['weight'])
                ws.cell(ri, 6, r['unit'])
                ws.cell(ri, 7, r['date'])
                ws.cell(ri, 8, r['time'])

            for col in ws.columns:
                width = max(len(str(c.value or '')) for c in col) + 4
                ws.column_dimensions[col[0].column_letter].width = width

            wb.save(xlsx_path)
            saved.append(os.path.basename(xlsx_path))

        messagebox.showinfo("Exported", "Saved to Desktop:\n\n" + "\n".join(saved))


if __name__ == '__main__':
    root = tk.Tk()
    app  = ScaleApp(root)
    root.mainloop()
